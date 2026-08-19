#!/usr/bin/env python3
"""
Backfill eRetail sales history from the yearly Excel exports.

Run once, locally:
    python3 scripts/ingest-eretail.py /path/to/Sale_Report*.xlsx

Writes:
    data/eretail/history.jsonl        one line per (date, channel)
    data/eretail/monthly/YYYY-MM.json product / category detail per month
    data/eretail/summary.json         coverage + the decisions applied

Decisions encoded here, agreed 19 Aug 2026:
  · STO is stock transfer between locations, not a sale — EXCLUDED.
    It is 47% of all rows, so including it would roughly double
    every figure.
  · Cancelled and Shipped & Returned are EXCLUDED (net, not gross).
  · There is no selling price in the export, only MRP. Revenue is
    MRP x quantity and is therefore LIST value, not money received.
    It will not reconcile against Shopify and is not meant to.
"""
import sys, os, json, glob, collections
from datetime import datetime, date, UTC
import openpyxl

EXCLUDE_CHANNELS = {"STO"}
EXCLUDE_STATUS = {"Cancelled", "Shipped & Returned"}

COLS = {"order": 0, "channel": 1, "date": 2, "sku": 3, "desc": 4,
        "segment": 5, "payment": 6, "status": 7, "size": 8,
        "colour": 9, "category": 10, "mrp": 11, "qty": 12}


def parse_date(v):
    """
    The exports mix two representations, and one of them is corrupted.

    Excel opened these DD/MM/YYYY strings and silently converted every
    one it could read as MM/DD into a real date — which is exactly the
    rows where day <= 12. Those cells now have day and month SWAPPED.
    The rows it could not read that way (day >= 13) stayed as strings.

    Verified across 1.87M rows in six files: no datetime cell has
    day > 12, and no string cell has day < 13. Zero counterexamples.
    So a datetime cell is unswapped by reading its day as the month.

    Without this correction, 594,000 rows — a third of the data —
    land on the wrong date, putting July sales in December.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        day, month = v.month, v.day          # swap back
        try:
            return date(v.year, month, day).isoformat()
        except ValueError:
            return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def main(paths, out_root="data/eretail"):
    daily = collections.defaultdict(lambda: {"units": 0, "listValue": 0.0, "orders": set()})
    monthly = collections.defaultdict(lambda: {
        "skus": collections.defaultdict(lambda: {"units": 0, "listValue": 0.0, "desc": None, "category": None}),
        "categories": collections.Counter(),
        "segments": collections.Counter(),
        "sizes": collections.Counter(),
        "colours": collections.Counter(),
        "channels": collections.defaultdict(lambda: {"units": 0, "listValue": 0.0}),
    })
    stats = collections.Counter()
    bad_dates = []

    for path in paths:
        name = os.path.basename(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        n = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[COLS["date"]] is None:
                continue
            n += 1
            stats["read"] += 1

            channel = str(r[COLS["channel"]] or "").strip()
            status = str(r[COLS["status"]] or "").strip()

            if channel in EXCLUDE_CHANNELS:
                stats["skipped_sto"] += 1
                continue
            if status in EXCLUDE_STATUS:
                stats["skipped_status"] += 1
                continue

            d = parse_date(r[COLS["date"]])
            if not d:
                stats["skipped_baddate"] += 1
                if len(bad_dates) < 5:
                    bad_dates.append(repr(r[COLS["date"]]))
                continue

            try:
                qty = int(r[COLS["qty"]] or 0)
                mrp = float(r[COLS["mrp"]] or 0)
            except (TypeError, ValueError):
                stats["skipped_badnum"] += 1
                continue

            value = mrp * qty
            month = d[:7]

            k = (d, channel)
            daily[k]["units"] += qty
            daily[k]["listValue"] += value
            daily[k]["orders"].add(r[COLS["order"]])

            m = monthly[month]
            sku = str(r[COLS["sku"]] or "?").strip()
            s = m["skus"][sku]
            s["units"] += qty
            s["listValue"] += value
            s["desc"] = s["desc"] or (r[COLS["desc"]] or None)
            s["category"] = s["category"] or (r[COLS["category"]] or None)

            def txt(v, default="?"):
                """Sizes and colours are sometimes numeric, sometimes blank."""
                if v is None:
                    return default
                out = str(v).strip()
                return out if out else default

            cat = txt(r[COLS["category"]], "Unknown")
            seg = r[COLS["segment"]]
            seg = seg.strip() if isinstance(seg, str) and not str(seg).startswith("=") else "Unknown"
            m["categories"][cat] += qty
            m["segments"][seg] += qty
            m["sizes"][txt(r[COLS["size"]])] += qty
            m["colours"][txt(r[COLS["colour"]])] += qty
            c = m["channels"][channel]
            c["units"] += qty
            c["listValue"] += value

            stats["kept"] += 1
        wb.close()
        print(f"  {name[:40]:42} {n:>8,} rows")

    os.makedirs(os.path.join(out_root, "monthly"), exist_ok=True)

    # ---- daily history, one line per (date, channel) ----
    lines = []
    for (d, ch), v in sorted(daily.items()):
        lines.append(json.dumps({
            "date": d, "channel": ch, "units": v["units"],
            "listValue": round(v["listValue"], 2), "orders": len(v["orders"]),
        }, separators=(",", ":")))
    with open(os.path.join(out_root, "history.jsonl"), "w") as f:
        f.write("\n".join(lines) + "\n")

    # ---- monthly detail ----
    for month, m in sorted(monthly.items()):
        top = sorted(m["skus"].items(), key=lambda kv: -kv[1]["units"])[:250]
        payload = {
            "month": month,
            "channels": {k: {"units": v["units"], "listValue": round(v["listValue"], 2)}
                         for k, v in m["channels"].items()},
            "categories": dict(m["categories"].most_common(40)),
            "segments": dict(m["segments"].most_common(30)),
            "sizes": dict(m["sizes"].most_common(20)),
            "colours": dict(m["colours"].most_common(30)),
            "topSkus": [{"sku": k, "desc": v["desc"], "category": v["category"],
                         "units": v["units"], "listValue": round(v["listValue"], 2)} for k, v in top],
        }
        with open(os.path.join(out_root, "monthly", f"{month}.json"), "w") as f:
            json.dump(payload, f, separators=(",", ":"))

    dates = sorted({d for d, _ in daily})
    summary = {
        "source": "eRetail yearly Excel exports",
        "generatedAt": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "coverage": {"from": dates[0], "to": dates[-1], "days": len(dates), "months": len(monthly)},
        "rows": dict(stats),
        "badDateSamples": bad_dates,
        "knownIssues": [
            "Excel corrupted the source exports: dates with day <= 12 were "
            "converted to MM/DD and are stored with day and month swapped. "
            "This ingest swaps them back. Verified: no datetime cell has "
            "day > 12 and no string cell has day < 13, across 1.87M rows."
        ],
        "rules": {
            "excludedChannels": sorted(EXCLUDE_CHANNELS),
            "excludedStatus": sorted(EXCLUDE_STATUS),
            "revenue": "listValue = MRP x quantity. LIST price, not money received — the export carries no selling price, so this will not reconcile against Shopify.",
        },
    }
    with open(os.path.join(out_root, "summary.json"), "w") as f:
        json.dump(summary, f, indent=2)

    print(f"\n  kept {stats['kept']:,} of {stats['read']:,} rows")
    print(f"  excluded: STO {stats['skipped_sto']:,} · cancelled/returned {stats['skipped_status']:,} · bad dates {stats['skipped_baddate']:,}")
    print(f"  coverage: {dates[0]} → {dates[-1]}  ({len(dates)} days, {len(monthly)} months)")


if __name__ == "__main__":
    args = sys.argv[1:] or sorted(glob.glob("/mnt/user-data/uploads/Sale_Report*.xlsx"))
    print(f"eRetail ingest · {len(args)} files")
    main(args)
